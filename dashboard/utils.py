import csv
import io
import os
from decimal import Decimal
from datetime import datetime
from django.http import HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders
from shop.models import Product

# For Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# For PDF export  
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.utils import ImageReader
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# For ODT export
try:
    from odf import text, style, table
    from odf.opendocument import OpenDocumentText
    from odf.style import Style, TextProperties, ParagraphProperties, TableColumnProperties
    from odf.table import Table as ODFTable, TableColumn, TableRow, TableCell
    from odf.text import P
    ODT_AVAILABLE = True
except ImportError:
    ODT_AVAILABLE = False


def export_products_csv(products, include_images=False):
    """Export products to CSV format"""
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="products_export_{timestamp}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    headers = ['ID', 'Name', 'Category', 'Description', 'Price (KSh)', 'Available', 'Orders Count']
    if include_images:
        headers.append('Image URL')
    
    writer.writerow(headers)
    
    # Write product data
    for product in products:
        row = [
            product.id,
            product.name,
            product.category.name,
            product.description,
            str(product.price),
            'Yes' if product.available else 'No',
            getattr(product, 'order_count', 0)
        ]
        
        if include_images:
            image_url = product.image.url if product.image else ''
            row.append(image_url)
        
        writer.writerow(row)
    
    return response


def export_products_excel(products, include_images=False):
    """Export products to Excel format"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    headers = ['ID', 'Name', 'Category', 'Description', 'Price (KSh)', 'Available', 'Orders Count']
    if include_images:
        headers.append('Image URL')
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write product data
    for row_num, product in enumerate(products, 2):
        worksheet.cell(row=row_num, column=1, value=product.id)
        worksheet.cell(row=row_num, column=2, value=product.name)
        worksheet.cell(row=row_num, column=3, value=product.category.name)
        worksheet.cell(row=row_num, column=4, value=product.description)
        worksheet.cell(row=row_num, column=5, value=float(product.price))
        worksheet.cell(row=row_num, column=6, value='Yes' if product.available else 'No')
        worksheet.cell(row=row_num, column=7, value=getattr(product, 'order_count', 0))
        
        if include_images:
            image_url = product.image.url if product.image else ''
            worksheet.cell(row=row_num, column=8, value=image_url)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="products_export_{timestamp}.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    return response


def export_products_pdf(products, include_images=False):
    """Export products to PDF format"""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install it with: pip install reportlab")
    
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="products_export_{timestamp}.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Add title
    title = Paragraph("Products Export Report", title_style)
    elements.append(title)
    
    # Add export info
    export_info = Paragraph(
        f"<b>Export Date:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>"
        f"<b>Total Products:</b> {len(products)}",
        styles['Normal']
    )
    elements.append(export_info)
    elements.append(Spacer(1, 20))
    
    # Prepare table data
    headers = ['ID', 'Name', 'Category', 'Price (KSh)', 'Available', 'Orders']
    table_data = [headers]
    
    for product in products:
        row = [
            str(product.id),
            product.name[:30] + '...' if len(product.name) > 30 else product.name,
            product.category.name,
            f"KSh {product.price}",
            'Yes' if product.available else 'No',
            str(getattr(product, 'order_count', 0))
        ]
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    return response


def export_products_odt(products, include_images=False):
    """Export products to ODT format"""
    if not ODT_AVAILABLE:
        raise ImportError("odfpy is required for ODT export. Install it with: pip install odfpy")
    
    # Create ODT document
    doc = OpenDocumentText()
    
    # Define styles
    bold_style = Style(name="Bold", family="text")
    bold_style.addElement(TextProperties(fontweight="bold"))
    doc.styles.addElement(bold_style)
    
    header_style = Style(name="Header", family="paragraph")
    header_style.addElement(ParagraphProperties(textalign="center"))
    doc.styles.addElement(header_style)
    
    # Add title
    title_para = P(stylename=header_style)
    title_para.addText("Products Export Report")
    doc.text.addElement(title_para)
    
    # Add export info
    info_para = P()
    info_para.addText(f"Export Date: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
    doc.text.addElement(info_para)
    
    info_para2 = P()
    info_para2.addText(f"Total Products: {len(products)}")
    doc.text.addElement(info_para2)
    
    # Add spacing
    doc.text.addElement(P())
    
    # Create table
    products_table = ODFTable(name="ProductsTable")
    
    # Add columns
    for i in range(6):  # ID, Name, Category, Price, Available, Orders
        products_table.addElement(TableColumn())
    
    # Add header row
    header_row = TableRow()
    headers = ['ID', 'Name', 'Category', 'Price (KSh)', 'Available', 'Orders']
    
    for header in headers:
        cell = TableCell()
        para = P(stylename=bold_style)
        para.addText(header)
        cell.addElement(para)
        header_row.addElement(cell)
    
    products_table.addElement(header_row)
    
    # Add product rows
    for product in products:
        row = TableRow()
        
        # Product data
        product_data = [
            str(product.id),
            product.name,
            product.category.name,
            f"KSh {product.price}",
            'Yes' if product.available else 'No',
            str(getattr(product, 'order_count', 0))
        ]
        
        for data in product_data:
            cell = TableCell()
            para = P()
            para.addText(data)
            cell.addElement(para)
            row.addElement(cell)
        
        products_table.addElement(row)
    
    doc.text.addElement(products_table)
    
    # Create response
    response = HttpResponse(content_type='application/vnd.oasis.opendocument.text')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="products_export_{timestamp}.odt"'
    
    # Save document to response
    doc.save(response)
    return response


def search_products(queryset, search_form):
    """Filter products based on search criteria"""
    if not search_form.is_valid():
        return queryset
    
    cleaned_data = search_form.cleaned_data
    
    # Search by name or description
    search_query = cleaned_data.get('search_query')
    if search_query:
        queryset = queryset.filter(
            name__icontains=search_query
        ).union(
            queryset.filter(description__icontains=search_query)
        )
    
    # Filter by category
    category = cleaned_data.get('category')
    if category:
        queryset = queryset.filter(category=category)
    
    # Filter by availability
    available = cleaned_data.get('available')
    if available == 'true':
        queryset = queryset.filter(available=True)
    elif available == 'false':
        queryset = queryset.filter(available=False)
    
    # Filter by price range
    price_min = cleaned_data.get('price_min')
    if price_min:
        queryset = queryset.filter(price__gte=price_min)
    
    price_max = cleaned_data.get('price_max')
    if price_max:
        queryset = queryset.filter(price__lte=price_max)
    
    return queryset


def get_export_requirements():
    """Get information about required packages for different export formats"""
    requirements = {
        'csv': {'available': True, 'package': 'Built-in'},
        'pdf': {'available': REPORTLAB_AVAILABLE, 'package': 'reportlab'},
        'xlsx': {'available': OPENPYXL_AVAILABLE, 'package': 'openpyxl'},
        'odt': {'available': ODT_AVAILABLE, 'package': 'odfpy'},
    }
    return requirements

